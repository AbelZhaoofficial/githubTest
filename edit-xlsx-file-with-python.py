import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}


print(product_list.max_row)
# range(75) == a list from 0 - 74
# range(2,75) == a list from 2 - 74
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    #It is grap and overwrite the row five, not create
    # we need the whole cell object to update a cell
    # cell 使用坐标锁定一个格子。这时supplier_name 里还是cell本身，不是cell里面的内容，需要后面加.value
    # print(supplier_name)

    # calculation for the number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)

        # get the value using the key name，.get do the same thing
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        # create and give value
        product_per_supplier[supplier_name] = 1

# print(product_per_supplier)

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:# new supplier
        total_value_per_supplier[supplier_name] = inventory * price

    # calculation of inventory number under 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    # this change the file ,but didn't save it.
    inventory_price.value = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
print(inventory_price)
inv_file.save("inventory_with_total_value.xlsx")